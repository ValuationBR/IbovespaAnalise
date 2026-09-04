"""
Gera o arquivo Excel final (saida/Ibovespa_Analise.xlsx) com:
 - Aba "Resumo": tabela com as 40 maiores empresas por valor de mercado,
   indicadores, preço-alvo e comentário automático
 - Uma aba por empresa com: gráfico de evolução do preço, próximos
   dividendos, os 20 indicadores (15 clássicos + 5 extras) e o valuation
   pelos 3 métodos.

Clique (ou passe o mouse) sobre o cabeçalho de qualquer indicador para
ver a fórmula usada e como interpretá-lo — isso é feito com comentários
de célula do Excel (openpyxl Comment).
"""
import os
from datetime import date
from decimal import Decimal
import psycopg2.extras
from db import conectar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

import config
from explicacoes import EXPLICACOES

FONTE = "Arial"
AZUL_ESCURO = "1F3864"
AZUL_CLARO = "DCE6F1"
VERDE = "C6EFCE"
VERMELHO = "FFC7CE"
BRANCO = "FFFFFF"

FONT_TITULO = Font(name=FONTE, size=16, bold=True, color=BRANCO)
FONT_CABECALHO = Font(name=FONTE, size=10, bold=True, color=BRANCO)
FONT_NORMAL = Font(name=FONTE, size=10)
FONT_LABEL = Font(name=FONTE, size=10, bold=True)
FILL_TITULO = PatternFill("solid", fgColor=AZUL_ESCURO)
FILL_CABECALHO = PatternFill("solid", fgColor=AZUL_ESCURO)
FILL_LABEL = PatternFill("solid", fgColor=AZUL_CLARO)
BORDA_FINA = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

AUTOR_COMENTARIO = "Painel Ibovespa"


def _num(v):
    return float(v) if isinstance(v, (int, float, Decimal)) else v


def _explicar(celula, titulo):
    texto = EXPLICACOES.get(titulo)
    if texto:
        celula.comment = Comment(texto, AUTOR_COMENTARIO, width=320, height=160)


def cabecalho_tabela(ws, linha, colunas, col_inicial=1):
    for i, titulo in enumerate(colunas):
        c = ws.cell(row=linha, column=col_inicial + i, value=titulo)
        c.font = FONT_CABECALHO
        c.fill = FILL_CABECALHO
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA_FINA
        _explicar(c, titulo)


def _observacoes_noticias(cur, ticker, limite=3):
    cur.execute(
        """SELECT titulo, fonte, data_publicacao, comentario_impacto FROM noticias
           WHERE ticker=%s ORDER BY data_publicacao DESC NULLS LAST LIMIT %s""",
        (ticker, limite),
    )
    linhas = cur.fetchall()
    if not linhas:
        return "Nenhuma notícia relevante coletada na última semana."
    partes = []
    for n in linhas:
        data_txt = n["data_publicacao"].strftime("%d/%m") if n["data_publicacao"] else "?"
        fonte_txt = f" ({n['fonte']})" if n["fonte"] else ""
        partes.append(f"[{data_txt}]{fonte_txt} {n['titulo']} — {n['comentario_impacto']}")
    return "\n".join(partes)


def aba_resumo(wb, cur):
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    ws["A1"].value = "Painel Ibovespa — 40 Maiores por Valor de Mercado"
    ws["A1"].font = Font(name=FONTE, size=18, bold=True, color=AZUL_ESCURO)
    ws["A2"].value = f"Atualizado em {date.today().strftime('%d/%m/%Y')} — clique nos cabeçalhos para ver a explicação de cada indicador"
    ws["A2"].font = Font(name=FONTE, size=10, italic=True, color="808080")

    colunas = [
        "Ticker", "Empresa", "Setor", "Preço Atual", "Valor de Mercado", "Variação 12m",
        "P/L", "P/VP", "ROE", "ROIC", "ROA", "Div. Yield", "Margem Líquida",
        "Dív.Líq/EBITDA", "EV/EBITDA", "Payout", "Beta", "Volatilidade Anual",
        "Alvo Graham", "Alvo Bazin", "Alvo DCF", "Alvo Médio", "Upside %",
        "Variação Alvo 12m", "Comentário", "Observações (notícias da semana)",
    ]
    linha_cab = 4
    cabecalho_tabela(ws, linha_cab, colunas)

    cur.execute(
        """SELECT e.ticker, e.nome, e.setor, i.preco_atual, i.market_cap, i.variacao_12m,
                  i.pl, i.pvp, i.roe, i.roic, i.roa, i.dividend_yield, i.margem_liquida,
                  i.divida_liquida_ebitda, i.ev_ebitda, i.payout, i.beta, i.volatilidade_anual,
                  v.preco_alvo_graham, v.preco_alvo_bazin, v.preco_alvo_dcf,
                  v.preco_alvo_medio, v.upside_medio_pct, v.variacao_alvo_medio_12m, i.comentario
           FROM empresas e
           LEFT JOIN indicadores i ON i.ticker = e.ticker
           LEFT JOIN valuation v ON v.ticker = e.ticker
           WHERE e.selecionada = TRUE
           ORDER BY i.market_cap DESC NULLS LAST"""
    )

    # colunas: índice (0-based) -> formato
    FORMATO_PRECO = '"R$"#,##0.00'
    FORMATO_MULTIPLO = "0.00x"
    FORMATO_PCT = "0.0%"
    FORMATOS = {
        3: FORMATO_PRECO, 4: '"R$"#,##0,,,"B"', 5: FORMATO_PCT,
        6: FORMATO_MULTIPLO, 7: FORMATO_MULTIPLO, 8: FORMATO_PCT, 9: FORMATO_PCT,
        10: FORMATO_PCT, 11: FORMATO_PCT, 12: FORMATO_PCT, 13: FORMATO_MULTIPLO,
        14: FORMATO_MULTIPLO, 15: FORMATO_PCT, 16: "0.00", 17: FORMATO_PCT,
        18: FORMATO_PRECO, 19: FORMATO_PRECO, 20: FORMATO_PRECO, 21: FORMATO_PRECO,
        22: FORMATO_PCT, 23: FORMATO_PCT,
    }

    linha = linha_cab + 1
    for row in cur.fetchall():
        observacoes = _observacoes_noticias(cur, row["ticker"])
        valores = [
            row["ticker"], row["nome"], row["setor"], row["preco_atual"], row["market_cap"],
            row["variacao_12m"], row["pl"], row["pvp"], row["roe"], row["roic"], row["roa"],
            row["dividend_yield"], row["margem_liquida"], row["divida_liquida_ebitda"],
            row["ev_ebitda"], row["payout"], row["beta"], row["volatilidade_anual"],
            row["preco_alvo_graham"], row["preco_alvo_bazin"], row["preco_alvo_dcf"],
            row["preco_alvo_medio"], row["upside_medio_pct"], row["variacao_alvo_medio_12m"],
            row["comentario"], observacoes,
        ]
        for i, v in enumerate(valores):
            col = i + 1
            c = ws.cell(row=linha, column=col, value=_num(v))
            c.font = FONT_NORMAL
            c.border = BORDA_FINA
            if i in FORMATOS:
                c.number_format = FORMATOS[i]
            if i in (24, 25):  # comentário, observações
                c.alignment = Alignment(wrap_text=True, vertical="top")
        # destaque de upside (verde se positivo, vermelho se negativo)
        cel_upside = ws.cell(row=linha, column=23)
        if isinstance(cel_upside.value, (int, float, Decimal)):
            cel_upside.fill = PatternFill("solid", fgColor=VERDE if cel_upside.value >= 0 else VERMELHO)
        linha += 1

    larguras = [9, 22, 16, 11, 13, 10, 8, 8, 9, 9, 9, 10, 12, 12, 10, 9, 8, 11, 11, 11, 11, 11, 10, 12, 45, 55]
    for i, largura in enumerate(larguras):
        ws.column_dimensions[get_column_letter(i + 1)].width = largura
    ws.freeze_panes = "D5"


def aba_empresa(wb, cur, ticker, nome):
    ws = wb.create_sheet(title=ticker[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"].value = f"{ticker} — {nome}"
    ws["A1"].font = Font(name=FONTE, size=16, bold=True, color=AZUL_ESCURO)

    # --- Histórico de preços (para o gráfico) ---
    cur.execute(
        "SELECT data_pregao, preco_fechamento FROM precos_diarios WHERE ticker=%s ORDER BY data_pregao",
        (ticker,),
    )
    precos = cur.fetchall()

    linha_inicio_precos = 3
    ws.cell(row=linha_inicio_precos, column=1, value="Data").font = FONT_CABECALHO
    ws.cell(row=linha_inicio_precos, column=2, value="Preço Fechamento").font = FONT_CABECALHO
    ws.cell(row=linha_inicio_precos, column=1).fill = FILL_CABECALHO
    ws.cell(row=linha_inicio_precos, column=2).fill = FILL_CABECALHO
    r = linha_inicio_precos + 1
    for p in precos:
        ws.cell(row=r, column=1, value=p["data_pregao"]).number_format = "dd/mm/yyyy"
        ws.cell(row=r, column=2, value=float(p["preco_fechamento"])).number_format = '"R$"#,##0.00'
        r += 1
    linha_fim_precos = r - 1

    if linha_fim_precos >= linha_inicio_precos + 1:
        chart = LineChart()
        chart.title = "Evolução do preço"
        chart.style = 2
        chart.y_axis.title = "R$"
        chart.x_axis.title = "Data"
        chart.height = 8
        chart.width = 18
        dados = Reference(ws, min_col=2, min_row=linha_inicio_precos, max_row=linha_fim_precos)
        categorias = Reference(ws, min_col=1, min_row=linha_inicio_precos + 1, max_row=linha_fim_precos)
        chart.add_data(dados, titles_from_data=True)
        chart.set_categories(categorias)
        chart.series[0].graphicalProperties.line.width = 20000
        ws.add_chart(chart, "D3")

    # --- Próximos dividendos ---
    linha_div_titulo = linha_fim_precos + 3
    ws.cell(row=linha_div_titulo, column=1, value="Próximos dividendos").font = Font(
        name=FONTE, size=12, bold=True, color=AZUL_ESCURO
    )
    cabecalho_tabela(ws, linha_div_titulo + 1, ["Tipo", "Valor por ação", "Data-com", "Pagamento"])
    cur.execute(
        """SELECT tipo, valor, data_com, data_pagamento FROM dividendos
           WHERE ticker=%s AND (data_pagamento >= CURRENT_DATE OR data_pagamento IS NULL)
           ORDER BY data_pagamento ASC NULLS LAST LIMIT 10""",
        (ticker,),
    )
    rr = linha_div_titulo + 2
    divs = cur.fetchall()
    if not divs:
        ws.cell(row=rr, column=1, value="Nenhum dividendo futuro anunciado no momento.").font = FONT_NORMAL
        rr += 1
    else:
        for d in divs:
            ws.cell(row=rr, column=1, value=d["tipo"]).font = FONT_NORMAL
            ws.cell(row=rr, column=2, value=float(d["valor"]) if d["valor"] else None).number_format = '"R$"#,##0.0000'
            ws.cell(row=rr, column=3, value=d["data_com"]).number_format = "dd/mm/yyyy"
            ws.cell(row=rr, column=4, value=d["data_pagamento"]).number_format = "dd/mm/yyyy"
            rr += 1

    # --- Indicadores (15 clássicos + 5 extras) ---
    linha_ind_titulo = rr + 2
    ws.cell(row=linha_ind_titulo, column=1, value="Indicadores (clique no nome para a explicação)").font = Font(
        name=FONTE, size=12, bold=True, color=AZUL_ESCURO
    )
    cur.execute("SELECT * FROM indicadores WHERE ticker=%s", (ticker,))
    ind = cur.fetchone() or {}
    lista_ind = [
        ("P/L", ind.get("pl"), "0.00x"), ("P/VP", ind.get("pvp"), "0.00x"),
        ("ROE", ind.get("roe"), "0.0%"), ("ROIC", ind.get("roic"), "0.0%"),
        ("ROA", ind.get("roa"), "0.0%"), ("Div. Yield", ind.get("dividend_yield"), "0.0%"),
        ("Margem Bruta", ind.get("margem_bruta"), "0.0%"), ("Margem Líquida", ind.get("margem_liquida"), "0.0%"),
        ("Margem EBITDA", ind.get("margem_ebitda"), "0.0%"), ("Liq. Corrente", ind.get("liquidez_corrente"), "0.00"),
        ("Dív.Líq/EBITDA", ind.get("divida_liquida_ebitda"), "0.00x"), ("EV/EBITDA", ind.get("ev_ebitda"), "0.00x"),
        ("EV/Receita", ind.get("ev_receita"), "0.00x"), ("Payout", ind.get("payout"), "0.0%"),
        ("Giro do Ativo", ind.get("giro_ativo"), "0.00"),
        ("Variação 12m", ind.get("variacao_12m"), "0.0%"),
        ("Correl. Carteira", ind.get("correlacao_carteira"), "0.00"),
        ("Correl. Setor", ind.get("correlacao_setor"), "0.00"),
        ("Volatilidade Anual", ind.get("volatilidade_anual"), "0.0%"),
        ("Beta", ind.get("beta"), "0.00"),
        ("CAGR Receita", ind.get("cagr_receita"), "0.0%"),
    ]
    r0 = linha_ind_titulo + 1
    linhas_por_coluna = 11
    for i, (nome_ind, valor, fmt) in enumerate(lista_ind):
        linha = r0 + (i % linhas_por_coluna)
        col = 1 if i < linhas_por_coluna else 3
        c_label = ws.cell(row=linha, column=col, value=nome_ind)
        c_label.font = FONT_LABEL
        c_label.fill = FILL_LABEL
        _explicar(c_label, nome_ind)
        cel = ws.cell(row=linha, column=col + 1, value=float(valor) if valor is not None else None)
        cel.number_format = fmt
        cel.font = FONT_NORMAL

    # --- Valuation ---
    linha_val_titulo = r0 + linhas_por_coluna + 2
    ws.cell(row=linha_val_titulo, column=1, value="Valuation — preço-alvo por método").font = Font(
        name=FONTE, size=12, bold=True, color=AZUL_ESCURO
    )
    cur.execute("SELECT * FROM valuation WHERE ticker=%s", (ticker,))
    val = cur.fetchone() or {}
    cabecalho_tabela(ws, linha_val_titulo + 1, ["Método", "Preço-alvo", "Upside vs. preço atual"])
    metodos = [
        ("Alvo Graham", val.get("preco_alvo_graham")),
        ("Alvo Bazin", val.get("preco_alvo_bazin")),
        ("Alvo DCF", val.get("preco_alvo_dcf")),
        ("Alvo Médio", val.get("preco_alvo_medio")),
    ]
    preco_atual = float(val.get("preco_atual")) if val.get("preco_atual") else None
    rv = linha_val_titulo + 2
    for nome_m, preco in metodos:
        preco = float(preco) if preco is not None else None
        c_nome = ws.cell(row=rv, column=1, value=nome_m)
        c_nome.font = FONT_NORMAL
        _explicar(c_nome, nome_m)
        c_preco = ws.cell(row=rv, column=2, value=preco)
        c_preco.number_format = '"R$"#,##0.00'
        if preco and preco_atual:
            upside = (preco / preco_atual) - 1
            c_up = ws.cell(row=rv, column=3, value=upside)
            c_up.number_format = "0.0%"
            c_up.fill = PatternFill("solid", fgColor=VERDE if upside >= 0 else VERMELHO)
        rv += 1

    # --- Comentário automático ---
    linha_comentario = rv + 2
    c_titulo = ws.cell(row=linha_comentario, column=1, value="Comentário")
    c_titulo.font = Font(name=FONTE, size=12, bold=True, color=AZUL_ESCURO)
    _explicar(c_titulo, "Comentário")
    ws.merge_cells(start_row=linha_comentario + 1, start_column=1, end_row=linha_comentario + 3, end_column=4)
    c_texto = ws.cell(row=linha_comentario + 1, column=1, value=ind.get("comentario") or "")
    c_texto.alignment = Alignment(wrap_text=True, vertical="top")
    c_texto.font = FONT_NORMAL

    for col, largura in zip("ABCD", [26, 18, 16, 16]):
        ws.column_dimensions[col].width = largura


def main():
    os.makedirs(config.PASTA_SAIDA, exist_ok=True)
    conn = conectar(cursor_factory=psycopg2.extras.RealDictCursor)
    cur = conn.cursor()

    wb = Workbook()
    aba_resumo(wb, cur)

    cur.execute("SELECT ticker, nome FROM empresas WHERE selecionada = TRUE ORDER BY ticker")
    empresas = cur.fetchall()
    for row in empresas:
        aba_empresa(wb, cur, row["ticker"], row["nome"])

    caminho = os.path.join(config.PASTA_SAIDA, "Ibovespa_Analise.xlsx")
    wb.save(caminho)
    cur.close()
    conn.close()
    print(f"Excel gerado em: {os.path.abspath(caminho)}")


if __name__ == "__main__":
    main()
